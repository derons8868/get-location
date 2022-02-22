from tkinter import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_all_elements_located
from selenium.common.exceptions import NoSuchElementException
import time
import string
import os 
import openpyxl
import sys
from openpyxl import Workbook, load_workbook


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

driver = webdriver.Edge(resource_path('./driver/msedgedriver.exe'))


def getmap():
    global driver
    #driver = webdriver.Chrome()
    driver.get("https://www.google.com/maps")
    
    
def getlocation():
    
    url1=driver.current_url
    lat, longitute = url1.split("!3d")[1].split("!4d");
    name=driver.find_element(By.XPATH, "/html/body/div[3]/div[9]/div[8]/div/div[1]/div/div/div[2]/div[1]/div[1]/div[1]/h1/span[1]")
    window.clipboard_clear()
    window.clipboard_append(name.text+" " + lat+" "+ longitute)
    namelabel1.configure(text=name.text)
    coordinatelable2.configure(text=lat)
    coordinatelable3.configure(text=longitute)
    return lat, longitute, name

def saveexcel():
    url1=driver.current_url
    lat, longitute = url1.split("!3d")[1].split("!4d");
    name=driver.find_element(By.XPATH, "/html/body/div[3]/div[9]/div[8]/div/div[1]/div/div/div[2]/div[1]/div[1]/div[1]/h1/span[1]")
    
    wb=openpyxl.load_workbook("coordinate.xlsx")
    ws=wb.worksheets[0]
    
    

    ws.append([lat,longitute,name.text])
    wb.save("coordinate.xlsx")
def createexcel():
    wb=Workbook()
    ws=wb.active
    ws.title="Coordinate"
    wb.save("coordinate.xlsx")
    
window = Tk()
window.geometry("800x500") 

mapbtn = Button(window, text="Open Google Map",command=getmap)
copybtn=Button(window, text="Copy to clipboard",command=getlocation)
namelabel=Label(window,text="Location Name")
namelabel1=Label(window, text=":")
excelbtn=Button(window, text="Copy to Excel file",command=saveexcel)
excelbtn1=Button(window, text="Create Excel file",command=createexcel)
coordinatelable=Label(window,text="Location Lattitude")
coordinatelable1=Label(window,text="Location Longtitude")
coordinatelable2=Label(window,text=":")
coordinatelable3=Label(window,text=":")


 





namelabel.grid(row=0,column=0, pady=50)
namelabel1.grid(row=0,column=1, pady=50)
coordinatelable.grid(row=1, column=0, pady=50)
coordinatelable1.grid(row=2, column=0, pady=50)
coordinatelable2.grid(row=1, column=1, pady=50)
coordinatelable3.grid(row=2, column=1, pady=50)
copybtn.grid(row=3,column=3,pady=10, padx=50)
mapbtn.grid(row=3, column=2)
excelbtn.grid(row=3, column=4)
excelbtn1.grid(row=3, column=1,padx=50)







window.mainloop()
